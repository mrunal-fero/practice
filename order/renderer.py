import json, openpyxl, io
from drf_excel.renderers import XLSXRenderer
from drf_excel.utilities import XLSXStyle, get_attribute
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.views import SheetView


class CustomXLSXRenderer(XLSXRenderer):
    def render(self, data, accepted_media_type=None, renderer_context=None):
        """
        Render `data` into XLSX workbook, returns a workbook.
        """
        if not self._check_validation_data(data):
            return json.dumps(data)

        if data is None:
            return bytes()

        wb = openpyxl.Workbook()
        self.ws = wb.active

        results = data["results"] if "results" in data else data

        drf_view = renderer_context.get("view")

        serializer = drf_view.serializer_class()
        if not serializer:
            raise ValueError("serializer_class missing!")

        header = get_attribute(drf_view, "header", {})
        use_header = header and header.get("use_header", True)
        self.ws.title = header.get("tab_title", "Report")
        header_title = header.get("header_title", "Report")

        column_header = getattr(drf_view, "column_header", {})
        column_count = 0
        row_count = 1
        if use_header:
            row_count += 1
        column_titles = column_header.get("titles", [])

        if len(results):
            use_labels = getattr(drf_view, "xlsx_use_labels", True)

            self.ignore_headers = getattr(drf_view, "xlsx_ignore_headers", [])

            self.boolean_display = getattr(drf_view, "xlsx_boolean_labels", None)

            self.column_data_styles = get_attribute(
                drf_view, "column_data_styles", dict()
            )

            self.custom_cols = getattr(drf_view, "xlsx_custom_cols", dict())

            self.custom_mappings = getattr(drf_view, "xlsx_custom_mappings", dict())

            self.fields_dict = self._serializer_fields(serializer)

            xlsx_header_dict = self._flatten_serializer_keys(
                serializer, use_labels=use_labels
            )
            if results and "custom_fields" in xlsx_header_dict:
                xlsx_header_dict.pop("custom_fields", None)
                for key, value in results[0].items():
                    if key not in xlsx_header_dict:
                        xlsx_header_dict[key] = key.title()

            if self.custom_cols:
                custom_header_dict = {
                    key: self.custom_cols[key].get("label", None) or key
                    for key in self.custom_cols.keys()
                }
                self.combined_header_dict = dict(
                    list(xlsx_header_dict.items()) + list(custom_header_dict.items())
                )
            else:
                self.combined_header_dict = xlsx_header_dict

            for column_name, column_label in self.combined_header_dict.items():
                if column_name == "row_color":
                    continue
                column_count += 1
                if column_count > len(column_titles):
                    column_name_display = column_label
                else:
                    column_name_display = column_titles[column_count - 1]

                header_cell = self.ws.cell(
                    row=row_count, column=column_count, value=column_name_display
                )
                # set_cell_style(header_cell, column_header_style)
            self.ws.row_dimensions[row_count].height = column_header.get("height", 45)

        # Set the header row
        if use_header:
            last_col_letter = "G"
            if column_count:
                last_col_letter = get_column_letter(column_count)
            self.ws.merge_cells("A1:{}1".format(last_col_letter))

            cell = self.ws.cell(row=1, column=1, value=header_title)
            # set_cell_style(cell, header_style)
            self.ws.row_dimensions[1].height = header.get("height", 45)

        # Set column width
        column_width = column_header.get("column_width", 20)
        if isinstance(column_width, list):
            for i, width in enumerate(column_width):
                col_letter = get_column_letter(i + 1)
                self.ws.column_dimensions[col_letter].width = width
        else:
            for ws_column in range(1, column_count + 1):
                col_letter = get_column_letter(ws_column)
                self.ws.column_dimensions[col_letter].width = column_width

        # Make body
        body = get_attribute(drf_view, "body", {})
        self.body_style = (
            XLSXStyle(body.get("style")) if body and "style" in body else None
        )
        if isinstance(results, dict):
            self._make_body(body, results, row_count)
        elif isinstance(results, list):
            for row in results:
                self._make_body(body, row, row_count)
                row_count += 1

        self.sheet_view_options = get_attribute(drf_view, "sheet_view_options", dict())
        self.ws.views.sheetView[0] = SheetView(**self.sheet_view_options)
        self.ws.freeze_panes = "A2"

        virtual_workbook = io.BytesIO()
        wb.save(virtual_workbook)
        return virtual_workbook.getvalue()